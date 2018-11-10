from openpyxl import load_workbook
from openpyxl import Workbook
path='/Users/ziliang.wu/Documents/data/'
file=path+'客服数据库总表3000总数据.xlsx'
newFile=path+'客服数据库总表3000总数据_new.xlsx'
newSimpleFile=path+'simple_new.xlsx'
wb=load_workbook(file)
wbNew = Workbook()
sheetNew = wbNew.active #新建一个excel文档并默认获取第一个工作表
sheetNew.title='合并数据'
wbsSimpleNew = Workbook()
sheetSimpleNew = wbsSimpleNew.active
sheetSimpleNew.title='问题和答案'
# 迭代输出
sheetnames=wb.sheetnames
for sheetname in sheetnames:
    sheet=wb[sheetname]
    temp=None
    for row in sheet.rows:
        array = []
        for cell in row:
            array.append(cell.value)  # 每行的单元格整合到一个数组，再整行添加到新的excel文档
        sheetNew.append(array)
        if temp!=row[1].value :
            arraySimle=[row[1].value,row[4].value]
            sheetSimpleNew.append(arraySimle)
        temp=row[1].value

wbNew.save(newFile)
wbsSimpleNew.save(newSimpleFile)