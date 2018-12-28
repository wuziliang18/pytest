from openpyxl import load_workbook
import csv
# path='/Users/ziliang.wu/Documents/data/'
path = 'D:/wuzl/data/客服问题/'
file = path + '111.xlsx'
newFile = path + '111_new.csv'
newSimpleFile = path + 'simple_new.csv'
wb = load_workbook(file)
csvNewFile = open(newFile, 'w', encoding='utf-8')
csvSimpleFile = open(newSimpleFile, 'w', encoding='utf-8')
csvNewFileWriter = csv.writer(csvNewFile)
csvSimpleFileWriter = csv.writer(csvSimpleFile)
# 迭代输出
sheetnames = wb.sheetnames
for sheetname in sheetnames:
    sheet = wb[sheetname]
    temp = None
    for row in sheet.rows:
        array = []
        for cell in row:
            array.append(cell.value)  # 每行的单元格整合到一个数组，再整行添加到新的excel文档
        csvNewFileWriter.writerow(array)
        if temp != row[1].value and row[1].value != None :
            arraySimle = [row[1].value + "`", row[4].value]
            csvSimpleFileWriter.writerow(arraySimle)
        temp = row[1].value
        

    
csvNewFile.close()
csvSimpleFile.close()
